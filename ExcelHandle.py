import datetime
import openpyxl
from openpyxl.styles import Font


class ExcelHandle(object):

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.wb.create_sheet(index=0, title='游戏猎人')
        self.sheet = self.wb['游戏猎人']
        self.fontObj = Font(name='Ebrima', size=14, bold=True, italic=False)

    def create_excel(self):
        self.sheet['A1'] = '发布时间'
        self.sheet['B1'] = '论坛ID'
        self.sheet['C1'] = '帖子标题'
        self.sheet['D1'] = '链接'
        self.sheet['E1'] = '阅读量'
        self.sheet['F1'] = '评论量'
        self.sheet['G1'] = '盖章'
        for x in range(ord('A'), ord('G') + 1):
            self.sheet[chr(x) + '1'].font = self.fontObj

    def add_excle(self, index, value):
        self.sheet[index] = value

    def save_excel(self, filePath):
        self.wb.save(filePath + '\\游戏猎人用户组数据.xlsx')

    def setFont(self, index, font):
        self.sheet[index].font = font;

    def merge_cell(self, interval):
        self.sheet.merge_cells(interval)

    def close_excel(self):
        self.wb.close()
